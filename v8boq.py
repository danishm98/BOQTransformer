import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import streamlit as st
import pandas as pd
from pandas import read_excel
import openpyxl
from IPython.display import display
import ipywidgets as widgets
import io
import os

st.title("BOQ Processing Workflow")

#uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

def master(uploaded_file):
    #if uploaded_file is not None:  
    excel_file = uploaded_file
    file_path = load_workbook(excel_file, data_only=True)

    
    
    # Input file path
    #file_path = 'Preprocessed_RLB - Amaala - Stage 2 - Hardscape BoQ Package (IFC cut-off 07.12.2023) RLB ASSESMENT.xlsx'
    #output_file_path = 'output_file_bill_Amaala__Stage 2 - Hardscape BoQ Package (IFC cut-off 07.12.2023) RLB ASSESMENT_____Part_2_test,,,,,,,,,,,.xlsx'
    
    # Extract filename without extension
    #base_filename = os.path.basename(file_path)
    #filename_without_extension = os.path.splitext(base_filename)[0][:25]  # Ensure the sheet name is within 31 characters
    
    # Create a new workbook and set the sheet name for the output
    wb = Workbook()
    ws = wb.active
    #ws.title = filename_without_extension
    
    # Apply column width settings
    ws.column_dimensions['A'].width = 50  # Set ITEM column width
    ws.column_dimensions['B'].width = 50  # Set DESCRIPTION column width
    ws.column_dimensions['C'].width = 20  # Set UNIT column width
    ws.column_dimensions['D'].width = 20  # Set QTY column width
    ws.column_dimensions['E'].width = 20  # Set RATE column width
    ws.column_dimensions['F'].width = 30  # Set AMOUNT column width
    
    # Apply text wrapping to all rows and columns
    wrap_text = Alignment(wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = wrap_text
    
    # Add headers to the first row
    
    


    # --------------------------------------------------------------------------------------------------------------------------------import pandas as pd

    # Unit conversion dictionary
    unit_conversion = {
        'KG': 'kg',
        'M': 'm',
        'M2': 'm²',
        'Item': 'item',
        'M³': 'm³',
        'M²': 'm²',
        'No': 'no.',
        'Kg': 'kg',
        'M3': 'm³'
    }
    
    def format_amount(amount):
        if isinstance(amount, float):
            return float(f"{amount:.2f}")
        return amount
    
    previous = False
    ignore_rows = False
    
    discard_keywords = [
        "addition for profit",  # Test case, but may need to include
        "cont'n..",
        "cont'n.",
        "cont'n",
        "cont…",
        "page total",
        "division",
        "summary",
        "page",
        "carried",
        "section",
        "tenderer",
        "tenderer's",
        "........."  # This is excluded but it may contain values in unit
    ]
    blank_row_counter = 0
    
    # Load the entire Excel file
    xls = pd.ExcelFile(file_path, engine='openpyxl')
    
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
        print(f"Processing Sheet: {sheet_name}")
        description_row_found = False
        colvalue = -1  
        description_row_index = -1

        sheet = xls.book[sheet_name]
        if sheet.sheet_state == 'hidden':
            continue

    
        # Find the header row
        #for index, row in df.iterrows():
        #    description_value = row[1]
        #    unit_value = str(row[2]).strip()
        #    if unit_value in unit_conversion:
        #        # Your existing logic here
        #        pass
    
    
        # Find the header row
        for index, row in df.iterrows():
            item_value = str(row[0])
            description_value = row[1]
            qty_value = row[3]
            unit_value = str(row[2]).strip()
            rate_value = row[4]
            amount_value = row[5]
            
            if unit_value in unit_conversion:
                unit_value = unit_conversion[unit_value]
            
            if pd.notna(item_value) and item_value.strip().lower() == "item" and \
               pd.notna(description_value) and description_value.strip().lower() == "description" and \
               pd.notna(qty_value) and qty_value.strip().lower() == "qty" and \
               pd.notna(unit_value) and unit_value.strip().lower() == "unit" and \
               pd.notna(rate_value) and rate_value.strip().lower() == "rate" and \
               pd.notna(amount_value) and amount_value.strip().lower() == "amount":
                description_row_found = True
                description_row_index = index
                break
    
    
        if not description_row_found:
            print(f"Error: Header row not found in sheet {sheet_name}.")
            continue
    
        # Find the AMOUNT column
        colvalue = -1
        for col_idx in range(len(df.columns) - 1, 1, -1):
            if pd.notna(df.iloc[description_row_index, col_idx]) and "amount" in str(df.iloc[description_row_index, col_idx]).lower():
                colvalue = col_idx
                break
    
        if colvalue == -1:
            print(f"Error: No AMOUNT column found for sheet {sheet_name}.")
            continue
    
        accumulated_description = ""
        last_bold_description = ""
        subheading_flag = False
    
        previous_item_index = None  # Initialize previous item index variable
    
        # Process rows after the header
        for index, row in df.iloc[description_row_index + 1:].iterrows():
            print(f"processing row:P{index}")
            if row.isna().all():
                blank_row_counter += 1
                if blank_row_counter >= 200:
                    print(f"Encountered 200 continuous blank rows in sheet {sheet_name}. Moving to next sheet.")
                    break
                continue
            else:
                blank_row_counter = 0
    
            
            description_value = row[1]
            if pd.isna(description_value):
                description_value = ""
            elif not isinstance(description_value, str):
                description_value = str(description_value)
    
            if any(keyword in description_value.strip().lower() for keyword in discard_keywords):
                continue  # Skip the rest of the processing for this row
            
            prefix_value = str(row[0]).strip() if pd.notna(row[0]) else ""
            qty_value = str(row[3]).strip() if pd.notna(row[0]) else ""
            amount_value = str(row[5]).strip() if pd.notna(row[5]) else ""
            rate_value = str(row[4]).strip() if pd.notna(row[4]) else ""
    
            if prefix_value.lower() == "excluded" or qty_value.lower() == "excluded" or amount_value.lower() == "excluded" or rate_value.lower() == "excluded" or description_value.strip().lower() == "excluded": # also, "included" and "by others"
                continue
            
            # Main Main headings (Division...)
            
            if description_value.isupper() and prefix_value != "" :
    
                accumulated_description = ""
                subheading_flag = False
                continue
    
            
            # Handle rows with prefix in format "x.y"
            if re.match(r'^\d+\.\d+$', prefix_value): # or is: (bold and underlined)
                accumulated_description =  description_value.strip()
                subheading_flag = False
                continue
    
            # Handle rows that contain only bold descriptions -- which are extended subheadings
            cell = sheet.cell(row=index + 1, column=2)
            if cell is not None and cell.font is not None:
    
    
                cell_bold_description = cell.font.bold
    
            if description_value.strip() != "" and (prefix_value == "" or pd.isna(prefix_value)) and (amount_value == "" or pd.isna(amount_value)) and cell_bold_description:
                if subheading_flag:
                    accumulated_description = accumulated_description.replace(last_bold_description, "").strip()
                last_bold_description = description_value.strip()
                accumulated_description += " " + last_bold_description
                subheading_flag = True
                
                continue
    
            # Handle actual line items
            if prefix_value !="" and qty_value != "" and amount_value != "" and rate_value != "":
                item = f"{description_value.strip()}"
                unit_value = str(row[2]).strip()
                if unit_value in unit_conversion:
                    unit_value = unit_conversion[unit_value]
                    
                qty_value = str(row[3])
                
                rate_value = str(row[4])
                
                
                amount_value = format_amount(row[5])
                ws.append([item, accumulated_description.strip(), unit_value, qty_value, rate_value, amount_value])
                previous_item_index = ws.max_row - 1  # Store the index of the previous item row for future reference
                previous_item_prefix = prefix_value   # Store the prefix of the previous item row for future reference
    
            # Handle rows with non-bold descriptions but blank everything else
            if description_value.strip() != "" and (prefix_value == "" or pd.isna(prefix_value)) and (amount_value == "" or pd.isna(amount_value)) and not cell_bold_description:
                if previous_item_index is not None:
                    ws.cell(row=previous_item_index + 1, column=1).value += f" {description_value.strip()}"
                continue
    
    # Apply text wrapping (probably not needed again)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):  # Columns 1 and 2 (ITEM, DESCRIPTION) and all others
        for cell in row:
            cell.alignment = wrap_text  # Enable text wrapping (again)
    
    
   
       
    
    # Assuming ws is your worksheet
    
    headers = ["Item", "Description", "Unit", "Qty", "Rate", "Amount"]
    #ws.append(headers)
    # Assign the headers to the first row
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Convert the data range to an actual Excel table
    table_range = f"A1:F{ws.max_row}"
    table = Table(displayName="Table1", ref=table_range)
    
    # Add a default style with banded rows
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    
    ws.add_table(table)

    # ----------------------------------------------------------------------------FINISH TRAINING THE MODEL ON NEW DATA 
    # ----------------------------------------------------------------------------INCLUDE THE PATTERN SUPERSET LOGIC ENCODED WITHIN THE MODEL
    # ----------------------------------------------------------------------------NOW LOAD THE XGBOOST + PATTERN HYBRID ML MODEL TO ASSIGN NRM + CATEGORY
    # ----------------------------------------------------------------------------ASSIGN THE NRM + CATEGORIES
    # ----------------------------------------------------------------------------PROMPT TO ENTER REST OF METADATA
    # ----------------------------------------------------------------------------FINAL FORMATTED FILE WHICH IS READY FOR POWER BI

    output_file_path = 'download test.xlsx'
    # Save the final formatted Excel file
    wb.save(output_file_path)
    print(f"Output saved to {output_file_path}")
    #updated_excel_path = "updated_excel.xlsx"
    
    return output_file_path
    #else:
    #    st.info("Please upload the BOQ Excel file to proceed.")


st.markdown("""
1. This workflow auto-processes a standard BOQ Excel file into a formatted table, and assigns a predicted NRM (and Category) against each line item.
2. Preprocessing requirements: Ideally, the file name should be under 50 characters.
3. Assumptions:
4. Instructions: 
5. If you face any issues or have questions, reach out to: Danish Memon
""")
excel_file = st.file_uploader("Select BOQ Excel File", type=["xlsx"])

if excel_file:
    # Read Excel file directly from the uploaded file
    excel_data = io.BytesIO(excel_file.getbuffer())
    
    updated_excel_path = master(excel_data)
    
    # Load the updated excel
    with open(updated_excel_path, "rb") as f:
        output_excel = io.BytesIO(f.read())
    
    st.success(f"File updated successfully! Download below")
    st.download_button(
        label="Download Updated Excel",
        data=output_excel,
        file_name="Updated formatted BOQ with NRM predictions.xlsx"
    )


